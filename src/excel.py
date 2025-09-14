#!/usr/bin/env python
# working with excel sheets

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell  # for type checker
from openpyxl.worksheet.worksheet import Worksheet  # for type checker


def find_date_row(sheet: Worksheet) -> int:
    """
    Determine the row number (row numbers count upwards from 1) with the dates.
    If we cannot figure this out return 0 to indicate a problem.

    MIN_DATE_COLUMNS_IN_ROW is the minimum number of columns next to each other
    in a row with cells having date type values for us to determine that the row
    is the one with dates. For the term roster this number is one half of a pay
    period, which is 7 days.
    """
    MIN_DATE_COLUMNS_IN_ROW = 7
    date_row = 0
    for row in sheet.iter_rows():
        date_count = 0
        for cell in row:
            assert isinstance(cell, Cell)  # Dbg
            if cell.is_date:
                date_count += 1
                if date_count == MIN_DATE_COLUMNS_IN_ROW and cell.row:
                    date_row = cell.row
                    break
    return date_row


# Find which column in the roster has the names
def find_name_column(sheet: Worksheet) -> str | None:
    """
    Determine which column has the names of the people rostered.
    REQUIRED_MATCH_COUNT is the number of matches for names to decide that the
    column contains names.
    """
    REQUIRED_MATCH_COUNT = 6
    regex = re.compile(r"\b[A-Z][a-z'-]+\s+[A-Z][a-z'-]+\b")
    column_letter = None
    for column in sheet.iter_cols():
        match_count = 0
        for cell in column:
            if cell.data_type == "s":
                if regex.match(cell.value):
                    match_count += 1
            if match_count == REQUIRED_MATCH_COUNT:
                column_letter = cell.column_letter
                break
    return column_letter


# Filter out list of possible names leaving only the full names
def filter_names_dict(name_to_row: dict[str, int]) -> dict[str, int]:
    """
    Filter out names from a list of possible names. List is from keys of the
    input dictionary.

    Parameters:
        - names_to_row: dict (possible name -> row number in roster)
    Return value:
        - dict with same format as input dict
        - name -> row number in roster
        - if multiple names found in same row each are mapped to same
          row number

    Names have following recognizable properties:
    - Has First name and last name
    - May/may not have a middle name
    - Names are separated by one/more spaces
    - Each name (part) starts with an uppercase letter
    - Subsequent letters are lowercase or hyphen or single-quote
    - Letter following hyphen or single-quote is uppercase
    """
    NAME_PART_PATTERN = r"\b(?:[A-Z](?:[a-z]+|[-\'][A-Z][a-z]*)+)\b"
    FULLNAME_PATTERN = rf"(?:{NAME_PART_PATTERN}\s+){{1,}}{NAME_PART_PATTERN}"
    regex = re.compile(FULLNAME_PATTERN)
    results: dict[str, int] = {}
    for string, row in name_to_row.items():
        matches = regex.finditer(string)
        fullnames = [match.group() for match in matches]
        if len(fullnames) > 0:
            for fullname in fullnames:
                results[fullname] = row
    return results


if __name__ == "__main__":
    from time import perf_counter

    t_start = perf_counter()
    t_load_times = 0
    ROSTER_DIR = "data"
    results: dict[str, dict[str, str | int | None]] = dict()
    roster_dir = Path(ROSTER_DIR)
    for i, roster_file in enumerate(roster_dir.glob("*.xlsx"), 1):
        print(f"Loading roster #{i}: {roster_file.name}")
        t0 = perf_counter()
        wb = load_workbook(roster_file)
        delta = perf_counter() - t0
        t_load_times += delta
        print(f"Done in {delta:.3f} seconds")
        print("------------------------------------------------------------")
        ws = wb.active
        if not ws:
            sys.stderr.write(f"Couldn't find active sheet of {roster_file.name}\n")
            continue
        date_row = find_date_row(ws)
        name_column = find_name_column(ws)
        results.setdefault(roster_file.name, {})["date_row"] = (
            date_row if date_row else None
        )
        results.setdefault(roster_file.name, {})["name_column"] = (
            name_column if name_column else None
        )

    t_end = perf_counter()
    max_len = max(len(k) for k in results)
    print(f"{'Roster File':^{max_len}s}\t{'Name Column':^11s}\t{'Date Row':^8s}")
    for filename, result in sorted(results.items()):
        print(
            f"{filename:>{max_len}}\t{result['name_column']:>11s}\t{result['date_row']:>8d}"
        )

    print(f"\nProgram took {t_end - t_start - t_load_times:.3f} seconds")
