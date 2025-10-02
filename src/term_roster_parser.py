#!/usr/bin/env python
# working with excel sheets

import re
import sys
from pathlib import Path
from typing import final

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell  # for type checker
from openpyxl.worksheet.worksheet import Worksheet  # for type checker
from openpyxl.utils import get_column_letter


def find_date_row(sheet: Worksheet) -> int:
    """
    Determine the row number (row numbers count upwards from 1) with the dates.
    If we cannot figure this out return 0 to indicate a problem.

    MIN_DATE_COLUMNS_IN_ROW is the minimum number of columns next to each other
    in a row with cells having date type values for us to determine that the row
    is the one with dates. For the term roster this number is one half of a pay
    period, which is 7 days.
    """
    MIN_DATE_COLUMNS_IN_ROW = 7
    date_row = 0
    for row in sheet.iter_rows():
        date_count = 0
        for cell in row:
            assert isinstance(cell, Cell)  # Dbg
            if cell.is_date:
                date_count += 1
                if date_count == MIN_DATE_COLUMNS_IN_ROW and cell.row:
                    date_row = cell.row
                    break
    return date_row


# Find which column in the roster has the names
def find_name_column(sheet: Worksheet) -> str:
    """
    Determine which column has the names of the people rostered.
    REQUIRED_MATCH_COUNT is the number of matches for names to decide that the
    column contains names.
    """
    REQUIRED_MATCH_COUNT = 6
    regex = re.compile(r"\b[A-Z][a-z'-]+\s+[A-Z][a-z'-]+\b")
    column_letter = ""
    for column in sheet.iter_cols():
        match_count = 0
        for cell in column:
            if cell.data_type == "s":
                if regex.match(cell.value):
                    match_count += 1
            if match_count == REQUIRED_MATCH_COUNT:
                column_letter = cell.column_letter
                break
    return column_letter


# Filter out list of possible names leaving only the full names
def filter_names_dict(name_to_row: dict[str, int]) -> dict[str, int]:
    """
    Filter out names from a list of possible names. List is from keys of the
    input dictionary.

    Parameters:
        - names_to_row: dict (possible name -> row number in roster)
    Return value:
        - dict with same format as input dict
        - name -> row number in roster
        - if multiple names found in same row each are mapped to same
          row number

    Names have following recognizable properties:
    - Has First name and last name
    - May/may not have a middle name
    - Names are separated by one/more spaces
    - Each name (part) starts with an uppercase letter
    - Subsequent letters are lowercase or hyphen or single-quote
    - Letter following hyphen or single-quote is uppercase
    """
    NAME_PART_PATTERN = r"\b(?:[A-Z](?:[a-z]+|[-\'][A-Z][a-z]*)+)\b"
    FULLNAME_PATTERN = rf"(?:{NAME_PART_PATTERN}\s+){{1,}}{NAME_PART_PATTERN}"
    regex = re.compile(FULLNAME_PATTERN)
    results: dict[str, int] = {}
    for string, row in name_to_row.items():
        matches = regex.finditer(string)
        fullnames = [match.group() for match in matches]
        if len(fullnames) > 0:
            for fullname in fullnames:
                results[fullname] = row
    return results


# Class to parse given term roster (excel) file and encapsulate all of its data.
# User can use the parser object to get specific information about the roster.
@final
class TermRosterParser:
    """
    Parse the term roster and encapsulate its contained data.

    Attributes:
        roster(Worksheet)
        date_row(int)
        name_column(int)
        name_to_row(dict[str, int])
    """

    def __init__(self, roster_file: str | Path):
        if not isinstance(roster_file, (str, Path)):
            raise TypeError(f'"{roster_file}" is not a valid string or pathlib.Path')
        if isinstance(roster_file, Path):
            roster_path = roster_file
        else:
            roster_path = Path(roster_file)
        if not roster_path.exists():
            raise FileNotFoundError(f'Path "{roster_file}" does not exist')
        if not roster_path.is_file():
            raise ValueError(f'Path "{roster_file}" exists, but is not a file')

        workbook = load_workbook(roster_path, data_only=True)
        worksheet = workbook.active
        if not worksheet:
            raise ValueError(f'"{roster_path}" does not have an active worksheet')

        self.roster: Worksheet = worksheet
        date_row = self.find_date_row()
        name_column = self.find_name_column()
        if date_row == 0:
            raise ValueError("Could not find the date row in the roster")
        if name_column is None:
            raise ValueError("Could not find the column with names in the roster")

        # Dbg
        sys.stderr.write(f"Date row @{date_row} | Name column @{name_column}\n")

        self.date_row = date_row
        self.name_column = name_column

        # Get mapping of row header to row number (some row headers are names)
        name_column_letter = get_column_letter(self.name_column)
        row_header_to_number: dict[str, int] = {
            cell.value: cell.row
            for cell in self.roster[name_column_letter]
            if cell.data_type == "s"
        }

        # Dbg
        sys.stderr.write(
            f"row-header-to-number has {len(row_header_to_number)} headers\n"
        )

        # Filter-out all the row headers except ones that are (probably) names.
        # If a row header has more than one name, map each name to same row
        # number.
        self.name_to_row = filter_names_dict(row_header_to_number)

        # Dbg
        sys.stderr.write(f"name-to-row dict has {len(self.name_to_row)} names\n")

    def find_date_row(self) -> int:
        """
        Determine the row number (row numbers count upwards from 1) with the dates.
        If we cannot figure this out return 0 to indicate a problem.

        Constants:
            REQUIRED_MIN_DATE_COLUMNS_IN_ROW:
                minimum number of columns next to each other in a row with cells
                having date type values for us to determine that the row is the
                one with dates. For the term roster this number is one half of a
                pay period, which is 7 days.
        Errors:
            Because row numbers start from 1, returning 0 means we could not
            find a date row.
        """
        REQUIRED_MIN_DATE_COLUMNS_IN_ROW = 7
        for row_number, row in enumerate(self.roster.iter_rows(), start=1):
            total_dates_in_row = sum(cell.is_date for cell in row)
            if total_dates_in_row >= REQUIRED_MIN_DATE_COLUMNS_IN_ROW:
                return row_number
        return 0

    def find_name_column(self) -> int:
        """
        Determine which column has the names of the people rostered.
        REQUIRED_MATCH_COUNT is the number of matches for names to conclude that
        the column contains the names.
        Columns start from 1, so 0 indicates invalid column
        Errors:
            Return value `None` means we could not find the name column
        """
        REQUIRED_MATCH_COUNT = 6
        regex = re.compile(r"\b[A-Z][a-z'-]+\s+[A-Z][a-z'-]+\b")
        for column in self.roster.iter_cols():
            match_count = 0
            for cell in column:
                if isinstance(cell, MergedCell):  # ? just to silent type-checker
                    continue
                if isinstance(cell.value, str):
                    if regex.match(cell.value):
                        match_count += 1
                if match_count == REQUIRED_MATCH_COUNT:
                    return cell.column
        return 0


if __name__ == "__main__":
    from time import perf_counter

    t_start = perf_counter()
    t_load_times = 0
    ROSTER_DIR = "data"
    results: dict[str, dict[str, str | int | None]] = dict()
    roster_dir = Path(ROSTER_DIR)
    for i, roster_file in enumerate(roster_dir.glob("*.xlsx"), 1):
        print(f"Loading roster #{i}: {roster_file.name}")
        t0 = perf_counter()
        wb = load_workbook(roster_file)
        delta = perf_counter() - t0
        t_load_times += delta
        print(f"Done in {delta:.3f} seconds")
        print("------------------------------------------------------------")
        ws = wb.active
        if not ws:
            sys.stderr.write(f"Couldn't find active sheet of {roster_file.name}\n")
            continue
        date_row = find_date_row(ws)
        name_column = find_name_column(ws)
        results.setdefault(roster_file.name, {})["date_row"] = (
            date_row if date_row else None
        )
        results.setdefault(roster_file.name, {})["name_column"] = (
            name_column if name_column else None
        )

    t_end = perf_counter()
    max_len = max(len(k) for k in results)
    print(f"{'Roster File':^{max_len}s}\t{'Name Column':^11s}\t{'Date Row':^8s}")
    for filename, result in sorted(results.items()):
        print(
            f"{filename:>{max_len}}\t{result['name_column']:>11s}\t{result['date_row']:>8d}"
        )

    print(f"\nProgram took {t_end - t_start - t_load_times:.3f} seconds")
